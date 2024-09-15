# [사용법]
# 1. 엑셀 파일에 유효성검사 필요한 인원 입력 후 저장 (반드시 값붙여넣기로 입력)
# 2. 해당 프로그램 오른쪽 상단 [▷] 버튼을 클릭하여 실행
# 3. 코드 실행 후 입력해야하는 총 인원수는 반드시 하단에 뜨는 질문의 옆을 클릭 후 입력 (그렇지 않으면 코드가 수정됨)
# 4. 오류가 발생할 경우 하단의 🗑(쓰레기통) 버튼 클릭

#License ver 1.0 (22.08.08)
#  ✓유효성검사 자동진행

#License ver 1.1 (22.08.09)
#  ✓식별번호 자동변환작업 (O->0, 0->O) 추가


from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl as op
import clipboard
import time
import pyautogui
import tkinter
from tkinter.filedialog import askopenfilename


max_number = pyautogui.prompt('총 인원수는 몇 명입니까?') #총 검사인원 입력
max_number = int(max_number)+2
start_time = time.time()
options = webdriver.ChromeOptions()
options.add_experimental_option("excludeSwitches", ["enable-logging"])
url = 'https://www.efine.go.kr/licen/truth/licenTruth.do?subMenuLv=010100'
driver = webdriver.Chrome(ChromeDriverManager().install())
driver.get(url)
driver.implicitly_wait(10)
root = tkinter.Tk()
root.withdraw()
filename = askopenfilename(parent=root)


wb = op.load_workbook(filename)
ws = wb.active  #활성화 되어있는 시트 설정(시트명 : "Sheet1")
ws_2 = wb['Sheet1']

for i in range(2,max_number): #반복 검사 시작
    #각 값별 변수 지정
    year = ws['A'+str(i)].value
    month = ws['B'+str(i)].value
    day = ws['C'+str(i)].value
    name = ws['D'+str(i)].value
    drive_license_1 = ws['E'+str(i)].value
    drive_license_2 = ws['F'+str(i)].value
    drive_license_3 = ws['G'+str(i)].value
    drive_license_4 = ws['H'+str(i)].value
    drive_license_5 = ws['I'+str(i)].value
    drive_license_5 = str(drive_license_5)
    count_O_drive_license_5 = drive_license_5.count('O')
    count_0_drive_license_5 = drive_license_5.count('0')
    
    
    #각 변수를 해당 위치에 입력
    driver.find_element(By.XPATH,'//*[@id="regYear"]').send_keys(year)
    driver.find_element(By.XPATH,'//*[@id="regMonth"]').send_keys(month)
    driver.find_element(By.XPATH,'//*[@id="regDate"]').send_keys(day)
    driver.find_element(By.XPATH,'//*[@id="name"]').send_keys(name)
    driver.find_element(By.XPATH,'//*[@id="licenNo0"]').send_keys(drive_license_1)
    driver.find_element(By.XPATH,'//*[@id="licenNo1"]').send_keys(drive_license_2)
    driver.find_element(By.XPATH,'//*[@id="licenNo2"]').send_keys(drive_license_3)
    driver.find_element(By.XPATH,'//*[@id="licenNo3"]').send_keys(drive_license_4)
    driver.find_element(By.XPATH,'//*[@id="ghostNo"]').send_keys(drive_license_5)

    #자동입력 방지문자 입력
    random_xpath = driver.find_element(By.XPATH,'//*[@id="btnSearch_msg1"]')
    random_text = random_xpath.get_attribute('Value')
    clipboard.copy(random_text)
    result = clipboard.paste()
    driver.find_element(By.XPATH,'//*[@id="btnSearch_msg0"]').send_keys(result)

    #조회
    check_btn = driver.find_element(By.XPATH,'//*[@id="btnSearch"]')
    check_btn.click()

    result_1 = driver.find_element(By.XPATH,'//*[@id="licen-truth"]/tbody/tr[1]/td/b/font').text

    #각 결과값 엑셀 시트에 입력
    if result_1 == '전산 자료와 일치 합니다.':
        result_2 = driver.find_element(By.XPATH,'//*[@id="licen-truth"]/tbody/tr[1]/td/b[2]/font').text
        if result_2 == '식별번호가 일치하지 않습니다.' and count_O_drive_license_5 > 0:
            new_O_drive_license_5 = drive_license_5.replace('O','0')
            driver.find_element(By.XPATH,'//*[@id="L_section"]/div[2]/ul/li[1]/a').click()
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH,'//*[@id="regYear"]').send_keys(year)
            driver.find_element(By.XPATH,'//*[@id="regMonth"]').send_keys(month)
            driver.find_element(By.XPATH,'//*[@id="regDate"]').send_keys(day)
            driver.find_element(By.XPATH,'//*[@id="name"]').send_keys(name)
            driver.find_element(By.XPATH,'//*[@id="licenNo0"]').send_keys(drive_license_1)
            driver.find_element(By.XPATH,'//*[@id="licenNo1"]').send_keys(drive_license_2)
            driver.find_element(By.XPATH,'//*[@id="licenNo2"]').send_keys(drive_license_3)
            driver.find_element(By.XPATH,'//*[@id="licenNo3"]').send_keys(drive_license_4)
            driver.find_element(By.XPATH,'//*[@id="ghostNo"]').send_keys(new_O_drive_license_5)
            #자동입력 방지문자 입력
            random_xpath = driver.find_element(By.XPATH,'//*[@id="btnSearch_msg1"]')
            random_text = random_xpath.get_attribute('Value')
            clipboard.copy(random_text)
            result = clipboard.paste()
            driver.find_element(By.XPATH,'//*[@id="btnSearch_msg0"]').send_keys(result)

            #조회
            check_btn = driver.find_element(By.XPATH,'//*[@id="btnSearch"]')
            check_btn.click()
            result_2 = driver.find_element(By.XPATH,'//*[@id="licen-truth"]/tbody/tr[1]/td/b[2]/font').text
            ws_2['J'+str(i)] = result_1
            ws_2['K'+str(i)] = result_2
            wb.save('C:\\Users\\user\\Desktop\\Python\\License.xlsx')
        elif result_2 == '식별번호가 일치하지 않습니다.' and count_0_drive_license_5 > 0:
            new_0_drive_license_5 = drive_license_5.replace('0','O')
            driver.find_element(By.XPATH,'//*[@id="L_section"]/div[2]/ul/li[1]/a').click()
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH,'//*[@id="regYear"]').send_keys(year)
            driver.find_element(By.XPATH,'//*[@id="regMonth"]').send_keys(month)
            driver.find_element(By.XPATH,'//*[@id="regDate"]').send_keys(day)
            driver.find_element(By.XPATH,'//*[@id="name"]').send_keys(name)
            driver.find_element(By.XPATH,'//*[@id="licenNo0"]').send_keys(drive_license_1)
            driver.find_element(By.XPATH,'//*[@id="licenNo1"]').send_keys(drive_license_2)
            driver.find_element(By.XPATH,'//*[@id="licenNo2"]').send_keys(drive_license_3)
            driver.find_element(By.XPATH,'//*[@id="licenNo3"]').send_keys(drive_license_4)
            driver.find_element(By.XPATH,'//*[@id="ghostNo"]').send_keys(new_0_drive_license_5)
            random_xpath = driver.find_element(By.XPATH,'//*[@id="btnSearch_msg1"]')
            random_text = random_xpath.get_attribute('Value')
            clipboard.copy(random_text)
            result = clipboard.paste()
            driver.find_element(By.XPATH,'//*[@id="btnSearch_msg0"]').send_keys(result)

            #조회
            check_btn = driver.find_element(By.XPATH,'//*[@id="btnSearch"]')
            check_btn.click()
            result_2 = driver.find_element(By.XPATH,'//*[@id="licen-truth"]/tbody/tr[1]/td/b[2]/font').text
            ws_2['J'+str(i)] = result_1
            ws_2['K'+str(i)] = result_2
            wb.save('C:\\Users\\user\\Desktop\\Python\\License.xlsx')
        else:
            ws_2['J'+str(i)] = result_1
            ws_2['K'+str(i)] = result_2
            wb.save('C:\\Users\\user\\Desktop\\Python\\License.xlsx')
    else:
        ws_2['J'+str(i)] = result_1
        ws_2['K'+str(i)] = result_1
        wb.save('C:\\Users\\user\\Desktop\\Python\\License.xlsx')

    #페이지 초기화
    driver.find_element(By.XPATH,'//*[@id="L_section"]/div[2]/ul/li[1]/a').click()
    time.sleep(0.4)


driver.quit() #크롬 닫기

#총 소요시간 측정
end_time = time.time()

print('코드 실행 시간: %20ds' % (end_time - start_time))